from scipy.integrate import quad
from openpyxl import Workbook
from openpyxl import load_workbook
import matplotlib.pyplot as plt
import csv
import pandas as pd
import os, logging
from os.path import dirname
from dotenv import load_dotenv  #libreria che semplifica l'uso di variabili d'ambiente, carica da pyenv le variabili
load_dotenv()

logging.getLogger('matplotlib').setLevel(logging.WARNING) 

def build_header_excel():
    headers = [
        'Run/Drivers', 'TravelTime (s)', 'LaneOffset (cm)',
        'Total Fitness Function', 'Best driver',
        'DesiredVelocity', 'DesiredAcceleration', 'DesiredDeceleration',
        'CurveBehavior', 'ObserveSpeedLimits', 'DistanceKeeping',
        'LaneKeeping', 'SpeedKeeping', 'LaneChangingDynamic',
        'UrgeToOvertake', 'RespondToTailgatingVehicles', 'ForesightDistance',
        'SteeringDistance', 'ObserveKeepRightRule', 'ConsiderEnvConditions',
        'UseOfIndicator', 'ReactionTime', 'ObeyTrafficSigns',
        'ObeyTrafficLights', 'ObeyTrafficRules', 'Swarm', 'RouteAdherence'] 

    return headers
         
def get_grafico(run_values, y_values, tipo_grafico, dir_path):
    
    cartella_radice = dir_path
    nome_sottocartella = 'Graphs'
    cartella = os.path.join(cartella_radice, nome_sottocartella)
    
    if not os.path.exists(cartella):
            os.makedirs(cartella)
    
    plt.plot(run_values, y_values)
    plt.title('Funzione Gaussiana')
    plt.xlabel('Run')
    plt.ylabel('Fitness function')
    plt.grid(True)

    nome_grafico = f'grafico_{tipo_grafico}.png'
    percorso_grafico = os.path.join(cartella, nome_grafico)
    plt.savefig(percorso_grafico)

    plt.close()

def get_path_excel():
    directory_path = '/home/udineoffice/Desktop/OptiDriverEvo_OnePlusOne/outputs'

    folders = [folder for folder in os.listdir(directory_path) if os.path.isdir(os.path.join(directory_path, folder))]

    folders.sort(key=lambda x: os.path.getctime(os.path.join(directory_path, x)), reverse=True)

    if folders:
        most_recent_folder = folders[0]
        
        file_to_store = os.path.join(directory_path, most_recent_folder, 'Results_of_simulations.xlsx')
        dir_path = os.path.join(directory_path, most_recent_folder)
        print(file_to_store)

    return file_to_store, dir_path

def dump_drivers(drivers, best_drivers):
    path_file_to_store, dir_path = get_path_excel()
    drivers_to_append = []
    for driver, best_driver in zip(drivers, best_drivers):
        row_driver_to_append = [driver.id_driver, driver.time_sim, driver.lane_offset, driver.total_fitness_function, best_driver.id_driver]
        for param in driver.parameters:
            row_driver_to_append.append(float(param))
        drivers_to_append.append(row_driver_to_append)       
          
    with open(os.path.join(dir_path,'driver_temp.csv'), 'a', newline='\n') as file_csv:
        writer = csv.writer(file_csv)
        writer.writerows(drivers_to_append)
        writer.writerow('\n')

def write_drivers_in_excel():
    path_file_to_store, dir_path = get_path_excel()
    with open(os.path.join(dir_path,'driver_temp.csv'), 'r') as file_csv:
        csv_reader = csv.reader(file_csv)
        data = []
        sheet_counter = 2
        with pd.ExcelWriter(path_file_to_store, engine='xlsxwriter') as writer:
            for row in csv_reader:
                if not any(row) or row == ['\n']:
                    if data:                        
                        df = pd.DataFrame(data)
                        df.to_excel(writer, sheet_name=f'Sim_{sheet_counter}', index=False, header=build_header_excel())
                        data = []
                        sheet_counter += 1
                else:
                    float_row = [float(column) for column in row]
                    data.append(float_row)

    #delete temp file
    os.remove(os.path.join(dir_path,'driver_temp.csv'))

'''def dump_drivers(drivers, best_drivers):

    run_values = []
    time_sim_values = []
    lane_offset_values = []
    fitness_function_values = []

    path_file_to_store, dir_path = get_path_csv()

    if os.path.exists(path_file_to_store):
        # Il file esiste, aprilo
        nuovo_file = load_workbook(path_file_to_store)
    else:
        # Il file non esiste, crea un nuovo file
        nuovo_file = Workbook()
        nuovo_file.save(path_file_to_store)

    if "Results of simulation 2" in nuovo_file.sheetnames:
        # Il foglio esiste già, crea un nuovo nome
        cont_of_sheet = 2
        while f"Results of simulation {cont_of_sheet}" in nuovo_file.sheetnames:
            cont_of_sheet += 1
        new_sheet_name = f"Results of simulation {cont_of_sheet}"
    else:
        # Il foglio non esiste, usa il titolo originale
        new_sheet_name = "Results of simulation 2"

    sheet = nuovo_file.create_sheet(title=new_sheet_name)
    print(sheet)

    build_header_file_csv(sheet)
    nuovo_file.save(path_file_to_store)
    num_of_run = 1

    for driver, best_driver in zip(drivers, best_drivers):
        
        run_values.append(num_of_run)
        time_sim = driver.time_sim
        time_sim_values.append(time_sim)
        lane_offset = driver.lane_offset
        lane_offset_values.append(lane_offset)
        fitness_function = driver.total_fitness_function
        fitness_function_values.append(fitness_function)

        #Scrivo i valori del mio driver
        sheet['A' + str(num_of_run + 1)] = driver.id_driver
        sheet['B' + str(num_of_run + 1)] = time_sim
        sheet['C' + str(num_of_run + 1)] = lane_offset
        sheet['D' + str(num_of_run + 1)] = fitness_function
        sheet['E' + str(num_of_run + 1)] = best_driver.id_driver

        colonna_parametri = 6
        #Scrivo i parametri del mio driver
        for params in driver.parameters:
            cella = sheet.cell(row=num_of_run+1, column=colonna_parametri)  
            cella.value = float(params)
            colonna_parametri += 1  # Passa alla colonna successiva per il prossimo parametro

        num_of_run += 1
        nuovo_file.save(path_file_to_store)
    
    nuovo_file.close()'''
   
    #get_grafico(valori_run, valori_time_sim, "time simulation", dir_path)
    #get_grafico(valori_run, valori_lane_offset, "lane offset", dir_path)
    #get_grafico(valori_run, valori_fitness_function, "fitness function totale", dir_path)
