from scipy.integrate import quad
from openpyxl import Workbook
from openpyxl import load_workbook
import matplotlib.pyplot as plt
import numpy as np, os, logging
from os.path import dirname
from dotenv import load_dotenv
load_dotenv()

logging.getLogger('matplotlib').setLevel(logging.WARNING) 
         
def get_grafico_pareto_front(total_population, dir_path):
    cartella_radice = dir_path    
    fitnesses = np.array([list(individual.fitness.values) for population in total_population for individual in population])
    plt.plot(fitnesses[:,0], fitnesses[:,1], "r.", label="Optimized")
    plt.title("Pareto Front")
    plt.xlabel("Time of simulation")
    plt.ylabel("Lane offset")
    plt.grid(True)

    nome_grafico = f'Pareto Front.png'
    percorso_grafico = os.path.join(cartella_radice, nome_grafico)
    plt.savefig(percorso_grafico)
    plt.close()

def get_path_csv():
    directory_path = '/home/udineoffice/Desktop/OptiDriverEvo_PopulationBased/outputs'
    folders = [folder for folder in os.listdir(directory_path) if os.path.isdir(os.path.join(directory_path, folder))]
    folders.sort(key=lambda x: os.path.getctime(os.path.join(directory_path, x)), reverse=True)

    if folders:
        most_recent_folder = folders[0]
        file_to_store = os.path.join(directory_path, most_recent_folder, 'Result_of_simulation.xlsx')
        dir_path = os.path.join(directory_path, most_recent_folder)

    return file_to_store, dir_path

def build_header_file_csv(sheet):
    headers = [
        'Generation', 'Best driver', 'TravelTime (s)', 
        'LaneOffset (cm)', 'MinTravelTime', 'MaxTravelTime', 'MinLaneOffset',
        'MaxLaneOffset', 'AvrTravelTime', 'AvrLaneOffset', 'StdTravelTime', 
        'StdLaneOffset', 'DesiredVelocity', 'DesiredAcceleration', 
        'DesiredDeceleration', 'CurveBehavior', 'ObserveSpeedLimits', 
        'DistanceKeeping', 'LaneKeeping', 'SpeedKeeping', 'LaneChangingDynamic',
        'UrgeToOvertake', 'RespondToTailgatingVehicles', 'ForesightDistance',
        'SteeringDistance', 'ObserveKeepRightRule', 'ConsiderEnvConditions',
        'UseOfIndicator', 'ReactionTime', 'ObeyTrafficSigns',
        'ObeyTrafficLights', 'ObeyTrafficRules', 'Swarm', 'RouteAdherence'] 

    for colonna, intestazione in enumerate(headers, start=1):
        cella = sheet.cell(row=1, column=colonna)
        cella.value = intestazione

def dump_best_populations(best_pops, logbook):

    path_file_to_store, dir_path = get_path_csv()
    nuovo_file = Workbook()
    sheet = nuovo_file.active
    sheet.title = "Results of total simulations"
    
    build_header_file_csv(sheet)
    nuovo_file.save(path_file_to_store)
    num_of_pop = 1
    num_of_driver = 1

    for pop in best_pops:        
        for driver in pop:       
            sheet['A' + str(num_of_driver + 1)] = num_of_pop
            sheet['B' + str(num_of_driver + 1)] = driver.id_driver
            sheet['C' + str(num_of_driver + 1)] = driver.fitness.values[0]
            sheet['D' + str(num_of_driver + 1)] = driver.fitness.values[1]

            nuovo_file.save(path_file_to_store) 

            colonna_parametri = 13
            for params in driver.parameters:
                cella = sheet.cell(row=num_of_driver+1, column=colonna_parametri)  
                cella.value = float(params)
                colonna_parametri += 1  
        
            num_of_driver += 1

        nuovo_file.save(path_file_to_store)   
        target_generation = num_of_pop - 1
        record_generation = None
        for generation_data in logbook:
            if generation_data['gen'] == target_generation:
                record_generation = generation_data
                break

        mins, maxs, avgs, stds = record_generation['min'], record_generation['max'], record_generation['avg'], record_generation['std']
        min_ts, max_ts, min_lo, max_lo, avg_ts, avg_lo, std_ts, std_lo = mins[0], maxs[0], mins[1], maxs[1], avgs[0], avgs[1], stds[0], stds[1]
        sheet['E' + str(num_of_driver)] = min_ts
        sheet['F' + str(num_of_driver)] = max_ts
        sheet['G' + str(num_of_driver)] = min_lo
        sheet['H' + str(num_of_driver)] = max_lo
        sheet['I' + str(num_of_driver)] = avg_ts
        sheet['J' + str(num_of_driver)] = avg_lo
        sheet['K' + str(num_of_driver)] = std_ts
        sheet['L' + str(num_of_driver)] = std_lo

        num_of_pop += 1
        nuovo_file.save(path_file_to_store)
            
    nuovo_file.save(path_file_to_store)
    nuovo_file.close()
    get_grafico_pareto_front(best_pops, dir_path)

