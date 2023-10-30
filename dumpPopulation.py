from scipy.integrate import quad
from openpyxl import Workbook
from openpyxl import load_workbook
import matplotlib.pyplot as plt
import numpy as np, os, logging, datetime
from os.path import dirname
from dotenv import load_dotenv  #libreria che semplifica l'uso di variabili d'ambiente, carica da pyenv le variabili
load_dotenv()

logging.getLogger('matplotlib').setLevel(logging.WARNING) 
         
def get_grafico_pareto_front(total_population, dir_path):
    cartella_radice = dir_path    
    #fitnesses_ini = np.array([list(individual.fitness.values) for population in total_population for individual in population])
    #fitnesses_ini = np.array([list(total_population[i].fitness.values) for i in range(len(total_population))])
    fitnesses = np.array([list(individual.fitness.values) for population in total_population for individual in population])
    #fitnesses = np.array([list(total_population[i].fitness.values) for i in range(len(total_population))])
    #plt.plot(fitnesses_ini[:,0], fitnesses_ini[:,1], "b.", label="Initial")
    plt.plot(fitnesses[:,0], fitnesses[:,1], "r.", label="Optimized")
    #plt.legend(loc="upper right", bbox_to_anchor=(1.15, 1))
    plt.title("Pareto Front")
    plt.xlabel("Time of simulation")
    plt.ylabel("Lane offset")
    plt.grid(True)

    nome_grafico = f'Pareto Front.png'
    percorso_grafico = os.path.join(cartella_radice, nome_grafico)
    plt.savefig(percorso_grafico)
    plt.close()

def get_grafico_features(logbook, dir_path):
    cartella_radice = dir_path
    gen = logbook.select("gen")
    fit_mins = logbook.chapters["fitness"].select("min")
    size_avgs = logbook.chapters["size"].select("avg")

    fig, ax1 = plt.subplots()
    line1 = ax1.plot(gen, fit_mins, "b-", label="Minimum Fitness")
    ax1.set_xlabel("Generation")
    ax1.set_ylabel("Fitness", color="b")
    for tl in ax1.get_yticklabels():
        tl.set_color("b")

    ax2 = ax1.twinx()
    line2 = ax2.plot(gen, size_avgs, "r-", label="Average Size")
    ax2.set_ylabel("Size", color="r")
    for tl in ax2.get_yticklabels():
        tl.set_color("r")

    lns = line1 + line2
    labs = [l.get_label() for l in lns]
    ax1.legend(lns, labs, loc="center right")

    nome_grafico = f'Min-Max-Average-Std.png'
    percorso_grafico = os.path.join(cartella_radice, nome_grafico)
    plt.savefig(percorso_grafico)
    plt.close()

def get_path_csv():
    # Specifica il percorso della directory principale
    directory_path = '/home/udineoffice/Desktop/OptiDriverEvo_PopulationBased/outputs'

    # Ottieni una lista di tutte le cartelle nella directory principale
    folders = [folder for folder in os.listdir(directory_path) if os.path.isdir(os.path.join(directory_path, folder))]

    # Ordina le cartelle per data di creazione, con la più recente in cima
    folders.sort(key=lambda x: os.path.getctime(os.path.join(directory_path, x)), reverse=True)

    # Prendi la prima cartella (la più recente)
    if folders:
        most_recent_folder = folders[0]
        # Costruisci il percorso completo del file da memorizzare
        file_to_store = os.path.join(directory_path, most_recent_folder, 'Result_of_simulation.xlsx')
        dir_path = os.path.join(directory_path, most_recent_folder)

    return file_to_store, dir_path

def build_header_file_csv(sheet):
    headers = [
        'Generation', 'Best driver', 'TravelTime (s)', 
        'LaneOffset (cm)', 'MinTravelTime', 'MaxTravelTime', 'MinLaneOffset',
        'MaxLaneOffset', 'AvrTravelTime', 'AvrLaneOffset', 'StdTravelTime', 
        'StdLaneOffset', 'DesiredVelocity', 'DesiredAcceleration', 
        'DesiredDeceleration', 'CurveBehavior', 'ObserveSpeedLimits', 
        'DistanceKeeping', 'LaneKeeping', 'SpeedKeeping', 'LaneChangingDynamic',
        'UrgeToOvertake', 'RespondToTailgatingVehicles', 'ForesightDistance',
        'SteeringDistance', 'ObserveKeepRightRule', 'ConsiderEnvConditions',
        'UseOfIndicator', 'ReactionTime', 'ObeyTrafficSigns',
        'ObeyTrafficLights', 'ObeyTrafficRules', 'Swarm', 'RouteAdherence'] 

    for colonna, intestazione in enumerate(headers, start=1):
        cella = sheet.cell(row=1, column=colonna)
        cella.value = intestazione

def dump_best_populations(best_pops, logbook):

    path_file_to_store, dir_path = get_path_csv()
    nuovo_file = Workbook()
    #nuovo_file = load_workbook(path_file_to_store)
    sheet = nuovo_file.active
    sheet.title = "Results of total simulations"
    
    build_header_file_csv(sheet)
    nuovo_file.save(path_file_to_store)
    num_of_pop = 1
    num_of_driver = 1

    for pop in best_pops:        
        for driver in pop:       
            sheet['A' + str(num_of_driver + 1)] = num_of_pop
            #Scrivo i valori del mio driver
            sheet['B' + str(num_of_driver + 1)] = driver.id_driver
            sheet['C' + str(num_of_driver + 1)] = driver.fitness.values[0]
            sheet['D' + str(num_of_driver + 1)] = driver.fitness.values[1]

            nuovo_file.save(path_file_to_store) 

            colonna_parametri = 13
            #Scrivo i parametri del mio driver
            for params in driver.parameters:
                cella = sheet.cell(row=num_of_driver+1, column=colonna_parametri)  
                cella.value = float(params)
                colonna_parametri += 1  # Passa alla colonna successiva per il prossimo parametro
        
            num_of_driver += 1

        nuovo_file.save(path_file_to_store)   
         
        # Recupera il record per la generazione desiderata
        target_generation = num_of_pop - 1 # Generazione desiderata
        # Trova il dizionario associato alla generazione desiderata
        record_generation = None
        for generation_data in logbook:
            if generation_data['gen'] == target_generation:
                record_generation = generation_data
                break

        mins, maxs, avgs, stds = record_generation['min'], record_generation['max'], record_generation['avg'], record_generation['std']
        min_ts, max_ts, min_lo, max_lo, avg_ts, avg_lo, std_ts, std_lo = mins[0], maxs[0], mins[1], maxs[1], avgs[0], avgs[1], stds[0], stds[1]
        sheet['E' + str(num_of_driver)] = min_ts
        sheet['F' + str(num_of_driver)] = max_ts
        sheet['G' + str(num_of_driver)] = min_lo
        sheet['H' + str(num_of_driver)] = max_lo
        sheet['I' + str(num_of_driver)] = avg_ts
        sheet['J' + str(num_of_driver)] = avg_lo
        sheet['K' + str(num_of_driver)] = std_ts
        sheet['L' + str(num_of_driver)] = std_lo

        num_of_pop += 1
        nuovo_file.save(path_file_to_store)
            
    nuovo_file.save(path_file_to_store)
    nuovo_file.close()
    get_grafico_pareto_front(best_pops, dir_path)
    #get_grafico_features(logbook, dir_path)

'''def dump_population(pops, logbook):
    path_file_to_store, dir_path = get_path_csv()

    if os.path.exists(path_file_to_store):
        # Il file esiste, aprilo
        nuovo_file = load_workbook(path_file_to_store)
    else:
        # Il file non esiste, crea un nuovo file
        nuovo_file = Workbook()
        nuovo_file.save(path_file_to_store)

    if "Population 0" in nuovo_file.sheetnames:
        # Il foglio esiste già, crea un nuovo nome
        cont_of_sheet = 0
        while f"Population {cont_of_sheet}" in nuovo_file.sheetnames:
            cont_of_sheet += 1
        new_sheet_name = f"Population {cont_of_sheet}"
    else:
        # Il foglio non esiste, usa il titolo originale
        new_sheet_name = "Population 0"

    sheet = nuovo_file.create_sheet(title=new_sheet_name)
    print(sheet)

    build_header_file_csv(sheet)
    nuovo_file.save(path_file_to_store)

    num_of_driver = 1

    for ind in pops:  # Partiamo dall'indice 1   
        sheet['A' + str(num_of_driver + 1)] = "Population"
        #Scrivo i valori del mio driver
        sheet['B' + str(num_of_driver + 1)] = ind.id_driver
        sheet['C' + str(num_of_driver + 1)] = ind.fitness.values[0]
        sheet['D' + str(num_of_driver + 1)] = ind.fitness.values[1]

        colonna_parametri = 6
        #Scrivo i parametri del mio driver
        for params in ind.parameters:
            cella = sheet.cell(row=num_of_driver+1, column=colonna_parametri)  
            cella.value = params
            colonna_parametri += 1  # Passa alla colonna successiva per il prossimo parametro
        
        num_of_driver += 1

    # Recupera il record per la generazione 1
    record_generation = logbook[1]

    # Accedi ai valori delle statistiche
    mins = record_generation['min']
    maxs = record_generation['max']
    avgs = record_generation['avg']
    stds = record_generation['std']
    min_ts = mins[0]
    max_ts = maxs[0]
    min_lo = mins[1]
    max_lo = maxs[1]
    avg_ts = avgs[0]
    std_ts = stds[0]
    avg_lo = avgs[1]
    std_lo = stds[1]
    sheet['A' + str(num_of_driver + 1)] = min_ts
    sheet['B' + str(num_of_driver + 1)] = max_ts
    sheet['C' + str(num_of_driver + 1)] = min_lo
    sheet['D' + str(num_of_driver + 1)] = max_lo
    sheet['E' + str(num_of_driver + 1)] = avg_ts
    sheet['F' + str(num_of_driver + 1)] = avg_lo
    sheet['G' + str(num_of_driver + 1)] = std_ts
    sheet['H' + str(num_of_driver + 1)] = std_lo
        
    nuovo_file.save(path_file_to_store)
    nuovo_file.close()
'''