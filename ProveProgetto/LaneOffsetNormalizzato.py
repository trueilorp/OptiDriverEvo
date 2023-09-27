import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
import numpy as np
import matplotlib.pyplot as plt
import pandas as pd
import os
import os, datetime
from os.path import dirname
from dotenv import load_dotenv  #libreria che semplifica l'uso di variabili d'ambiente, carica da pyenv le variabili
load_dotenv()

valoriYGrafico = []
valoriRun = []
valoriLaneOffset = []

def gaussiana(x, media, deviazione_std):
    return (1 / (deviazione_std * np.sqrt(2 * np.pi))) * np.exp(-(x - media)**2 / (2 * deviazione_std**2))

def doGrafico():
    
    cartella_radice = '/home/udineoffice/Desktop'
    nome_sottocartella = 'Graphs'
    cartella = os.path.join(cartella_radice, nome_sottocartella)

    if not os.path.exists(cartella):
            os.makedirs(cartella)    

    # Visualizza la funzione gaussiana
    plt.plot(valoriRun, valoriYGrafico)
    plt.title('Funzione Gaussiana')
    plt.xlabel('run')
    plt.ylabel('ff2')
    plt.grid(True)

    nome_grafico = f'grafico_LaneOffset.png'
    percorso_grafico = os.path.join(cartella, nome_grafico)
    plt.savefig(percorso_grafico)

    plt.close()


def normalizzaLO(valore, i):
    
    try:
    # Prova ad aprire il file Excel
        nuovo_file = load_workbook(filename='/home/udineoffice/Desktop/Graphs.xlsx')
        sheet_name = "LaneOffset"
        if sheet_name not in nuovo_file.sheetnames:
            # Crea un nuovo foglio di lavoro solo se non esiste
            sheet = nuovo_file.create_sheet("LaneOffset")
            print(f"Foglio di lavoro '{sheet_name}' creato.")
        sheet = nuovo_file.get_sheet_by_name('LaneOffset')
        print("File Excel esistente aperto in modalità di sola lettura.")
    except FileNotFoundError:
        # Se il file non esiste, crea un nuovo file Excel
        nuovo_file = Workbook()
        sheet = nuovo_file.create_sheet("LaneOffset")
        sheet = nuovo_file.get_sheet_by_name('LaneOffset')
        print("Nuovo file Excel creato.")

    #foglio = nuovo_file.active
    sheet['A1'] = 'Run'
    sheet['B1'] = 'laneOffset (cm)'
    sheet['C1'] = 'Average'
    sheet['D1'] = 'Deviazione Standard'
    sheet['E1'] = 'Distribuzione Gaussiana'
    sheet['F1'] = '1 - ff1'
    
    nuovo_file.save('/home/udineoffice/Desktop/Graphs.xlsx')

    colonne_fisse = ['laneOffset (cm)'] 
    colonnaRun = ['Run'] 
    # Carica il file Excel in un DataFrame
    df = pd.read_excel('/home/udineoffice/Desktop/Graphs.xlsx')
    # Assumi che i dati siano nella colonna "Dati"

    sheet['B' + str(i+1)] = valore
    x = sheet.cell(row=i+1, column=2).value
    valoriLaneOffset.append(x) #metto i valori del time sim in una lista 
    
    datiPerDSTD = np.array(valoriLaneOffset)

    #Scrivo il numero dalla run
    sheet['A' + str(i+1)] = i
    valoriRun.append(sheet.cell(row=i+1, column=1).value)

    #Calcola la media parziale
    media = round(np.mean(valoriLaneOffset),2)
    sheet['C' + str(i+1)] = media

    #Calcola la deviazione standard
    deviazione_std = round(np.std(datiPerDSTD),2)
    #deviazione_std = round(np.std(valoriTimeSim),2)
    sheet['D' + str(i+1)] = deviazione_std

    #Calcola la funzione gaussiana
    y = gaussiana(x, media, deviazione_std)
    sheet['E' + str(i+1)] = y

    #Calcola 1 - funzione gaussiana
    w = 1 - y
    sheet['F' + str(i+1)] = w
    valoriYGrafico.append(sheet.cell(row=i+1, column=6).value)

    print("MEDIA " + str(media))
    print("DEVIAZIONE STD " + str(deviazione_std))
    print("VALORE NORMALIZZATO " + str(y))

    nuovo_file.save('/home/udineoffice/Desktop/Graphs.xlsx')

    # chiudiamo il file
    nuovo_file.close()

    return w 

    #doGrafico(valoriRun, valoriYGrafico, i)






